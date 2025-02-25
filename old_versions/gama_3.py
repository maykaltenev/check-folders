import os
import hashlib
import argparse
import openpyxl
from openpyxl.styles import Font
from tqdm import tqdm


def get_all_files(folder):
    """Recursively get all file paths in a folder, relative to the base folder."""
    file_paths = []
    for root, _, files in os.walk(folder):
        for file in files:
            relative_path = os.path.relpath(os.path.join(root, file), folder)
            file_paths.append(relative_path)
    return set(file_paths)


def get_file_hash(filepath):
    """Generate SHA256 hash of a file."""
    hash_sha256 = hashlib.sha256()
    try:
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_sha256.update(chunk)
        return hash_sha256.hexdigest()
    except Exception as e:
        print(f"❌ Failed to read {filepath}: {e}")
        return None


def clean_sheet_name(name):
    """Remove invalid characters from an Excel sheet name."""
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
    for char in invalid_chars:
        name = name.replace(char, "-")
    return name[:31]  # Excel sheet name max length is 31 characters


def detect_changed_files(folder1, folder2, extra_files1, extra_files2):
    """
    Check for files that exist only in one folder but are actually the same content
    (i.e. changed file names). Returns:
      - changed_files: list of tuples (original, changed)
      - updated_extra1: extra files in folder1 with renamed ones removed
      - updated_extra2: extra files in folder2 with renamed ones removed
    """
    changed_files = []
    unmatched_extra_files1 = set(extra_files1)
    unmatched_extra_files2 = set(extra_files2)

    # Build a hash mapping for folder1 extra files.
    hash_to_files1 = {}
    for file in list(unmatched_extra_files1):
        file1_path = os.path.join(folder1, file)
        hash_val = get_file_hash(file1_path)
        if hash_val is not None:
            hash_to_files1.setdefault(hash_val, []).append(file)

    # Check folder2 extra files against folder1's hashes.
    for file in list(unmatched_extra_files2):
        file2_path = os.path.join(folder2, file)
        hash_val = get_file_hash(file2_path)
        if hash_val is not None and hash_val in hash_to_files1:
            # Assume one-to-one mapping: take the first match.
            orig_file = hash_to_files1[hash_val].pop(0)
            changed_files.append((orig_file, file))
            unmatched_extra_files1.discard(orig_file)
            unmatched_extra_files2.discard(file)
            # If no more files with this hash remain, remove the key.
            if not hash_to_files1[hash_val]:
                del hash_to_files1[hash_val]

    return list(changed_files), list(unmatched_extra_files1), list(unmatched_extra_files2)


def compare_folders(folder1, folder2):
    """Compare two folders recursively with a structured output."""
    print("\n🔍 Scanning folders...\n")
    files1 = get_all_files(folder1)
    files2 = get_all_files(folder2)

    only_in_folder1 = sorted(files1 - files2)
    only_in_folder2 = sorted(files2 - files1)
    common_files = sorted(files1 & files2)

    different_files = []
    identical_files = []
    failed_comparisons = []
    total_checked = 0

    print("\n🔄 Comparing common files...\n")
    for file in tqdm(common_files, desc="Comparing", unit="file"):
        file1_path = os.path.join(folder1, file)
        file2_path = os.path.join(folder2, file)

        hash1 = get_file_hash(file1_path)
        hash2 = get_file_hash(file2_path)

        if hash1 is None or hash2 is None:
            failed_comparisons.append(file)
        elif hash1 != hash2:
            different_files.append(file)
        else:
            identical_files.append(file)
        total_checked += 1

    # Detect changed file names among extra files.
    changed_files, only_in_folder1, only_in_folder2 = detect_changed_files(
        folder1, folder2, only_in_folder1, only_in_folder2)

    # Console output for better feedback
    print("\n📌 Comparison Summary:\n")
    print(f"🔍 Total files checked: {total_checked}")
    print(f"✅ Identical files: {len(identical_files)}")
    print(f"🔄 Different files: {len(different_files)}")
    print(f"📁 Only in {folder1}: {len(only_in_folder1)}")
    print(f"📁 Only in {folder2}: {len(only_in_folder2)}")
    print(f"❌ Failed comparisons: {len(failed_comparisons)}")

    if changed_files:
        print("\n📌 Changed File Names:")
        for original, changed in changed_files:
            print(f"  🔄 {original} ➝ {changed}")

    if only_in_folder1:
        print(f"\n📁 {folder1} has {len(only_in_folder1)} extra files:")
        for f in only_in_folder1:
            print(f"  ➤ {f}")

    if only_in_folder2:
        print(f"\n📁 {folder2} has {len(only_in_folder2)} extra files:")
        for f in only_in_folder2:
            print(f"  ➤ {f}")

    if different_files:
        print("\n🔄 Different files found:")
        for f in different_files:
            print(f"  ⚠️ {f}")

    if failed_comparisons:
        print("\n⚠️ Files that could not be read:")
        for f in failed_comparisons:
            print(f"  ❌ {f}")

    # Save results to Excel (including changed file names)
    save_to_excel(folder1, folder2, only_in_folder1, only_in_folder2,
                  different_files, identical_files, failed_comparisons,
                  changed_files, total_checked)


def save_to_excel(folder1, folder2, only_in_folder1, only_in_folder2,
                  different_files, identical_files, failed_comparisons,
                  changed_files, total_checked):
    """Save comparison results to an Excel file with detailed sheets."""
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = clean_sheet_name("Folder Comparison")

    # Formatting: Bold headers
    bold_font = Font(bold=True)

    # Main sheet headers
    headers = ["File Path", "Status", "Location"]
    ws.append(headers)
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header).font = bold_font

    # Populate rows for files in both folders.
    all_files = (
        [(f, "Extra", folder1) for f in only_in_folder1] +
        [(f, "Extra", folder2) for f in only_in_folder2] +
        [(f, "Different", "Both folders") for f in different_files] +
        [(f, "Identical", "Both folders") for f in identical_files] +
        [(f, "Failed", "Comparison failed") for f in failed_comparisons]
    )

    for file_path, status, location in all_files:
        ws.append([file_path, status, location])

    # Summary sheet
    summary_ws = workbook.create_sheet(title="Summary")
    summary_ws.append(["Category", "Count"])
    summary_ws.append(["Total files checked", total_checked])
    summary_ws.append(["Identical files", len(identical_files)])
    summary_ws.append(["Different files", len(different_files)])
    summary_ws.append(["Changed file names", len(changed_files)])
    summary_ws.append(["Only in " + folder1, len(only_in_folder1)])
    summary_ws.append(["Only in " + folder2, len(only_in_folder2)])
    summary_ws.append(["Failed comparisons", len(failed_comparisons)])

    for col in summary_ws.columns:
        max_length = max(len(str(cell.value))
                         if cell.value else 0 for cell in col)
        summary_ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Auto-adjust column widths in the main sheet
    for col in ws.columns:
        max_length = max(len(str(cell.value))
                         if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # New sheet for changed file names
    if changed_files:
        changed_ws = workbook.create_sheet(title="Changed File Names")
        changed_ws.append(["Original File Path", "Changed File Path"])
        for col_num in range(1, 3):
            changed_ws.cell(row=1, column=col_num).font = bold_font
        for original, changed in changed_files:
            changed_ws.append([original, changed])
        # Adjust column widths
        for col in changed_ws.columns:
            max_length = max(len(str(cell.value))
                             if cell.value else 0 for cell in col)
            changed_ws.column_dimensions[col[0]
                                         .column_letter].width = max_length + 2

    # Save file
    excel_filename = "folder_comparison.xlsx"
    workbook.save(excel_filename)
    print(f"\n📄 Comparison saved to: {excel_filename}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Compare two folders recursively and generate an Excel report.")
    parser.add_argument("folder1", nargs="?", help="Path to the first folder")
    parser.add_argument("folder2", nargs="?", help="Path to the second folder")
    args = parser.parse_args()

    # Ask for folder input if not provided in command line
    if not args.folder1:
        args.folder1 = input("📁 Enter the path for Folder 1: ").strip()
    if not args.folder2:
        args.folder2 = input("📁 Enter the path for Folder 2: ").strip()

    compare_folders(args.folder1, args.folder2)
