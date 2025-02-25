import os
import hashlib
import argparse
import openpyxl
from openpyxl.styles import Font
from tqdm import tqdm


def get_all_files(folder):
    """Recursively get all file paths in a folder, relative to the base folder."""
    file_paths = []
    for root, _, files in os.walk(folder):
        for file in files:
            relative_path = os.path.relpath(os.path.join(root, file), folder)
            file_paths.append(relative_path)
    return set(file_paths)


def get_file_hash(filepath):
    """Generate SHA256 hash of a file."""
    hash_sha256 = hashlib.sha256()
    try:
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_sha256.update(chunk)
        return hash_sha256.hexdigest()
    except Exception as e:
        print(f"❌ Failed to read {filepath}: {e}")
        return None


def clean_sheet_name(name):
    """Remove invalid characters from an Excel sheet name."""
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
    for char in invalid_chars:
        name = name.replace(char, "-")
    return name[:31]  # Excel sheet name max length is 31 characters


def compare_folders(folder1, folder2):
    """Compare two folders recursively with a progress bar and console feedback."""
    print("\n🔍 Scanning folders...\n")
    files1 = get_all_files(folder1)
    files2 = get_all_files(folder2)

    only_in_folder1 = sorted(files1 - files2)
    only_in_folder2 = sorted(files2 - files1)
    common_files = sorted(files1 & files2)

    different_files = []
    identical_files = []
    failed_comparisons = []
    total_checked = 0

    print("\n🔄 Comparing common files...\n")
    for file in tqdm(common_files, desc="Comparing", unit="file"):
        file1_path = os.path.join(folder1, file)
        file2_path = os.path.join(folder2, file)

        hash1 = get_file_hash(file1_path)
        hash2 = get_file_hash(file2_path)

        if hash1 is None or hash2 is None:
            failed_comparisons.append(file)
        elif hash1 != hash2:
            different_files.append(file)
        else:
            identical_files.append(file)
        total_checked += 1

    # Console output for better feedback
    print("\n📌 Comparison Summary:\n")
    print(f"🔍 Total files checked: {total_checked}")
    print(f"✅ Identical files: {len(identical_files)}")
    print(f"🔄 Different files: {len(different_files)}")
    print(f"📁 Only in {folder1}: {len(only_in_folder1)}")
    print(f"📁 Only in {folder2}: {len(only_in_folder2)}")
    print(f"❌ Failed comparisons: {len(failed_comparisons)}")

    if failed_comparisons:
        print("\n⚠️ Files that could not be read:")
        for f in failed_comparisons[:10]:
            print(f"  ❌ {f}")
        if len(failed_comparisons) > 10:
            print("  ... and more.")

    # Save to Excel
    save_to_excel(folder1, folder2, only_in_folder1, only_in_folder2,
                  different_files, identical_files, failed_comparisons, total_checked)


def save_to_excel(folder1, folder2, only_in_folder1, only_in_folder2, different_files, identical_files, failed_comparisons, total_checked):
    """Save comparison results to an Excel file with better feedback."""
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = clean_sheet_name("Folder Comparison")

    # Formatting: Bold headers
    bold_font = Font(bold=True)

    # Headers
    headers = ["File Path", "Status", "Location"]
    ws.append(headers)

    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header).font = bold_font

    # Populate rows
    all_files = (
        [(f, "Extra", folder1) for f in only_in_folder1] +
        [(f, "Extra", folder2) for f in only_in_folder2] +
        [(f, "Different", "Both folders") for f in different_files] +
        [(f, "Identical", "Both folders") for f in identical_files] +
        [(f, "Failed", "Comparison failed") for f in failed_comparisons]
    )

    for file_path, status, location in all_files:
        ws.append([file_path, status, location])

    # Summary sheet
    summary_ws = workbook.create_sheet(title="Summary")
    summary_ws.append(["Category", "Count"])
    summary_ws.append(["Total files checked", total_checked])
    summary_ws.append(["Identical files", len(identical_files)])
    summary_ws.append(["Different files", len(different_files)])
    summary_ws.append(["Only in " + folder1, len(only_in_folder1)])
    summary_ws.append(["Only in " + folder2, len(only_in_folder2)])
    summary_ws.append(["Failed comparisons", len(failed_comparisons)])

    for col in summary_ws.columns:
        max_length = max(len(str(cell.value))
                         if cell.value else 0 for cell in col)
        summary_ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value))
                         if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Save file
    excel_filename = "folder_comparison.xlsx"
    workbook.save(excel_filename)
    print(f"\n📄 Comparison saved to: {excel_filename}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Compare two folders recursively and generate an Excel report.")
    parser.add_argument("folder1", nargs="?", help="Path to the first folder")
    parser.add_argument("folder2", nargs="?", help="Path to the second folder")
    args = parser.parse_args()

    # Ask for folder input if not provided in command line
    if not args.folder1:
        args.folder1 = input("📁 Enter the path for Folder 1: ").strip()
    if not args.folder2:
        args.folder2 = input("📁 Enter the path for Folder 2: ").strip()

    compare_folders(args.folder1, args.folder2)
