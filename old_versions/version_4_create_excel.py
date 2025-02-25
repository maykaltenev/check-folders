import os
import hashlib
import argparse
import openpyxl


def get_all_files(folder):
    """Recursively get all file paths in a folder, relative to the base folder."""
    file_paths = []
    for root, _, files in os.walk(folder):
        for file in files:
            relative_path = os.path.relpath(os.path.join(root, file), folder)
            file_paths.append(relative_path)
    return set(file_paths)


def get_file_hash(filepath):
    """Generate SHA256 hash of a file."""
    hash_sha256 = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_sha256.update(chunk)
    return hash_sha256.hexdigest()


def compare_folders(folder1, folder2):
    """Compare two folders recursively and save results in an Excel file."""
    files1 = get_all_files(folder1)
    files2 = get_all_files(folder2)

    only_in_folder1 = files1 - files2
    only_in_folder2 = files2 - files1
    common_files = files1 & files2

    different_files = []
    identical_files = []

    for file in sorted(common_files):
        file1_path = os.path.join(folder1, file)
        file2_path = os.path.join(folder2, file)

        hash1 = get_file_hash(file1_path)
        hash2 = get_file_hash(file2_path)

        if hash1 != hash2:
            different_files.append(file)
        else:
            identical_files.append(file)

    # Log results to an Excel file
    save_to_excel(folder1, folder2, only_in_folder1,
                  only_in_folder2, different_files, identical_files)


def save_to_excel(folder1, folder2, missing1, missing2, different_files, identical_files):
    """Save the comparison results into an Excel file."""
    wb = openpyxl.Workbook()

    # Missing/Extra Files Sheet (fixed sheet name)
    ws1 = wb.active
    ws1.title = "Missing_Extra_Files"  # ✅ Fix here
    ws1.append(["File Path", "Status"])
    for file in sorted(missing1):
        ws1.append([os.path.join(folder1, file), "Only in Folder 1"])
    for file in sorted(missing2):
        ws1.append([os.path.join(folder2, file), "Only in Folder 2"])

    # Different Files Sheet
    ws2 = wb.create_sheet(title="Different_Files")  # ✅ Fix here
    ws2.append(["File Path"])
    for file in sorted(different_files):
        ws2.append([file])

    # Identical Files Sheet
    ws3 = wb.create_sheet(title="Identical_Files")  # ✅ Fix here
    ws3.append(["File Path"])
    for file in sorted(identical_files):
        ws3.append([file])

    # Save the file
    excel_filename = "comparison_results.xlsx"
    wb.save(excel_filename)
    print(f"\n✅ Results saved to {excel_filename}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Compare two folders recursively and log results to Excel.")
    parser.add_argument("folder1", help="Path to the first folder")
    parser.add_argument("folder2", help="Path to the second folder")
    args = parser.parse_args()

    compare_folders(args.folder1, args.folder2)
